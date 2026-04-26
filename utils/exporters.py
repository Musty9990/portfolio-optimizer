"""
Exporters Module
================
Export portfolio analysis results to Excel/CSV.
"""

from datetime import datetime
from io import BytesIO
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.optimizer import PortfolioResult


HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)


def export_to_excel(
    max_sharpe: PortfolioResult,
    min_vol: PortfolioResult,
    prices: pd.DataFrame,
    frontier: Optional[pd.DataFrame] = None,
) -> bytes:
    """
    Build a multi-sheet Excel report and return as bytes (for Streamlit download).

    Sheets:
        - Summary    : key metrics for both portfolios
        - Max Sharpe : weights detail
        - Min Vol    : weights detail
        - Prices     : raw price history
        - Frontier   : (optional) efficient frontier points
    """
    wb = Workbook()

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "PORTFOLIO OPTIMIZATION REPORT"
    ws["A1"].font = Font(bold=True, size=16, color="1F4E79")
    ws.merge_cells("A1:C1")

    ws["A3"] = "Generated:"
    ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws["A4"] = "Period:"
    ws["B4"] = f"{prices.index.min().date()} → {prices.index.max().date()}"
    ws["A5"] = "Assets:"
    ws["B5"] = ", ".join(prices.columns)

    headers = ["Metric", "Max Sharpe", "Min Volatility"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    rows = [
        ("Expected Return", f"{max_sharpe.expected_return*100:.2f}%",
         f"{min_vol.expected_return*100:.2f}%"),
        ("Volatility",      f"{max_sharpe.volatility*100:.2f}%",
         f"{min_vol.volatility*100:.2f}%"),
        ("Sharpe Ratio",    f"{max_sharpe.sharpe_ratio:.3f}",
         f"{min_vol.sharpe_ratio:.3f}"),
    ]
    for i, row in enumerate(rows, start=8):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)

    for col_idx in range(1, 4):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    # Per-portfolio weight sheets
    for name, result in [("Max Sharpe", max_sharpe), ("Min Vol", min_vol)]:
        sheet = wb.create_sheet(name)
        sheet["A1"] = f"{name} Portfolio Weights"
        sheet["A1"].font = Font(bold=True, size=14)
        sheet.merge_cells("A1:B1")

        for col, h in enumerate(["Ticker", "Weight (%)"], start=1):
            cell = sheet.cell(row=3, column=col, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        sorted_weights = sorted(result.weights.items(), key=lambda x: -x[1])
        for i, (ticker, w) in enumerate(sorted_weights, start=4):
            sheet.cell(row=i, column=1, value=ticker)
            sheet.cell(row=i, column=2, value=round(w * 100, 2))

        sheet.column_dimensions["A"].width = 18
        sheet.column_dimensions["B"].width = 14

    # Raw prices sheet
    prices_sheet = wb.create_sheet("Prices")
    prices_sheet.append(["Date"] + list(prices.columns))
    for date, row in prices.iterrows():
        prices_sheet.append([date.date()] + [round(v, 2) for v in row.values])
    for col_idx in range(1, len(prices.columns) + 2):
        prices_sheet.column_dimensions[get_column_letter(col_idx)].width = 14

    # Frontier sheet (optional)
    if frontier is not None and not frontier.empty:
        f_sheet = wb.create_sheet("Frontier")
        f_sheet.append(["Volatility (%)", "Return (%)"])
        for _, row in frontier.iterrows():
            f_sheet.append([round(row["volatility"] * 100, 4),
                            round(row["return"] * 100, 4)])
        f_sheet.column_dimensions["A"].width = 18
        f_sheet.column_dimensions["B"].width = 18

    # Return as bytes for download
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
